"""
convert_members_xlsx.py의 main() 전체 흐름을 실제 xlsx/json 파일(임시 디렉터리)로
돌려보는 통합 테스트. 단위 테스트(test_convert_members_xlsx.py)는 resolve_conflict()
같은 개별 함수만 검증하는데, 실제 엑셀 파일을 읽고 쓰는 openpyxl 연동과 baseline
스냅샷 파일이 회차를 거듭하며 제대로 쌓이는지는 여기서 확인한다.
"""

import json

from openpyxl import Workbook, load_workbook

import convert_members_xlsx as cmx

HEADER = ["이름", "SOOP ID", "ELO ID", "생년월일", "성별", "종족", "티어", "소속", "직책", "수정일"]


def _make_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "members"
    ws.append(HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _setup(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "members.xlsx"
    members_path = tmp_path / "members.json"
    baseline_path = tmp_path / "members_sync_baseline.json"
    monkeypatch.setattr(cmx, "XLSX_PATH", xlsx_path)
    monkeypatch.setattr(cmx, "MEMBERS_PATH", members_path)
    monkeypatch.setattr(cmx, "BASELINE_PATH", baseline_path)
    return xlsx_path, members_path, baseline_path


class TestMainEndToEnd:
    def test_first_run_populates_json_from_xlsx(self, tmp_path, monkeypatch):
        xlsx_path, members_path, baseline_path = _setup(tmp_path, monkeypatch)
        _make_xlsx(xlsx_path, [
            ["조기석", "sharpragu", 3, None, "남자", "테란", "갓", "JSA", None, None],
        ])

        cmx.main()

        assert members_path.exists()
        data = json.loads(members_path.read_text(encoding="utf-8"))
        assert len(data["members"]) == 1
        assert data["members"][0]["id"] == "sharpragu"
        assert data["members"][0]["nickname"] == "조기석"
        assert baseline_path.exists()

    def test_admin_edit_propagates_back_to_xlsx_on_next_run(self, tmp_path, monkeypatch):
        """admin.html에서 json을 직접 고친 상황을 재현 - 다음 실행에서 xlsx에도
        반영돼야 한다(양방향 동기화의 핵심)."""
        xlsx_path, members_path, baseline_path = _setup(tmp_path, monkeypatch)
        _make_xlsx(xlsx_path, [
            ["조기석", "sharpragu", 3, None, "남자", "테란", "갓", "JSA", None, None],
        ])
        cmx.main()  # 1차 실행 - baseline 생성

        # admin.html이 members.json의 team을 직접 고쳤다고 가정
        data = json.loads(members_path.read_text(encoding="utf-8"))
        data["members"][0]["team"] = "ADMIN수정팀"
        members_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        cmx.main()  # 2차 실행 - xlsx로 반영돼야 함

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb["members"]
        row = list(ws.iter_rows(min_row=2, values_only=True))[0]
        assert row[7] == "ADMIN수정팀"  # 소속 컬럼

    def test_new_json_only_member_gets_appended_to_xlsx(self, tmp_path, monkeypatch):
        xlsx_path, members_path, baseline_path = _setup(tmp_path, monkeypatch)
        _make_xlsx(xlsx_path, [
            ["조기석", "sharpragu", 3, None, "남자", "테란", "갓", "JSA", None, None],
        ])
        cmx.main()

        # sync_members.py가 새 멤버를 json에만 추가했다고 가정
        data = json.loads(members_path.read_text(encoding="utf-8"))
        data["members"].append({
            "id": "newbie", "elo_id": 999, "nickname": "새멤버", "birthdate": None,
            "gender": "m", "race": "저그", "tier": "0", "team": "FA", "role": "",
            "info_updated_at": None,
        })
        members_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        cmx.main()

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb["members"]
        ids = [r[1] for r in ws.iter_rows(min_row=2, values_only=True)]
        assert "newbie" in ids

    def test_second_run_with_no_changes_is_idempotent(self, tmp_path, monkeypatch):
        """아무것도 안 바뀐 상태로 두 번째 실행하면 json/xlsx 내용이 똑같이
        유지돼야 한다(불필요한 변경이 생기면 안 됨)."""
        xlsx_path, members_path, baseline_path = _setup(tmp_path, monkeypatch)
        _make_xlsx(xlsx_path, [
            ["조기석", "sharpragu", 3, None, "남자", "테란", "갓", "JSA", None, None],
        ])
        cmx.main()
        first_json = members_path.read_text(encoding="utf-8")

        cmx.main()  # 아무것도 안 바뀐 채로 재실행
        second_json = members_path.read_text(encoding="utf-8")

        assert first_json == second_json

    def test_missing_xlsx_does_not_crash(self, tmp_path, monkeypatch):
        xlsx_path, members_path, baseline_path = _setup(tmp_path, monkeypatch)
        # xlsx 파일을 아예 안 만듦
        cmx.main()  # 예외 없이 끝나야 함
        assert not members_path.exists()  # json도 안 건드려짐
