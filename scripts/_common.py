"""
ststats 프로젝트의 스크립트들이 공통으로 쓰는 유틸리티.

fetch_poonggo_data.py, fetch_eloboard_data.py, download_font.py, generate_pages.py
전부 여기서 ROOT 경로와 시간/숫자 관련 헬퍼를 가져다 쓴다. 이전에는 스크립트마다
kst_now()/_to_int()/DATETIME_FORMAT을 똑같이 복사해서 갖고 있었는데, 하나만
고치고 나머지를 깜빡하는 실수를 막기 위해 한 곳으로 모았다.

HTTP 재시도 로직도 마찬가지 이유로 여기 있다 - fetch_poonggo_data.py(urllib)와
fetch_eloboard_data.py(requests)가 각자 따로 재시도 루프를 갖고 있었는데(그나마
풍고 쪽은 재시도 사이에 대기(backoff)도 없어서 실패하자마자 곧바로 재요청하는
차이까지 있었다), fetch_json() 하나로 합쳐서 두 스크립트가 동일한 타임아웃/재시도
횟수/백오프 간격, 동일한 형태의 경고 로그를 쓰게 했다.
"""

import sys
import os
import json
import time
import calendar
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

# scripts/_common.py 기준으로 두 단계 위(레포 루트)
ROOT = Path(__file__).resolve().parent.parent

DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"

# 외부 사이트에 요청 보낼 때 공통으로 쓰는 User-Agent. 자동화된 요청임을 정직하게
# 밝힌다(사전에 정식 이용 허가를 받은 풍고 API 기준으로 정한 문구이며, 다른
# 스크립트에서도 동일하게 재사용한다).
USER_AGENT = "ststats-bot/1.0 (+https://ststats.github.io)"

# fetch_json()의 기본값. 두 fetch 스크립트가 동일한 값을 쓰도록 여기 한 곳에서만
# 관리한다 - 서버 쪽 사정(타임아웃 늘려야 한다든가)이 바뀌면 여기만 고치면 된다.
HTTP_TIMEOUT_SEC = 30
HTTP_MAX_RETRIES = 3
HTTP_RETRY_BACKOFF_SEC = 3


def kst_now() -> datetime:
    """한국 표준시(KST, UTC+9) 기준 현재 시각."""
    return datetime.now(timezone.utc) + timedelta(hours=9)


def to_int(value) -> int:
    """API 응답 등에서 온 값을 정수로 안전하게 변환. 콤마 섞인 문자열("1,234")이나
    None도 처리하고, 변환 불가능하면 0을 반환한다(예외로 전체 스크립트가 죽지 않도록)."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).replace(",", "").strip() or 0)
    except (ValueError, TypeError):
        return 0


def last_day_of_month(year: int, month: int) -> int:
    """그 해/달의 마지막 날짜(28~31)를 반환한다. fetch_poonggo_data.py와
    fetch_eloboard_data.py 둘 다 "달이 바뀐 첫 실행에서 지난달 마지막 날 아카이브를
    확정한다" 로직에서 이 계산이 필요해서 공용으로 뺐다."""
    return calendar.monthrange(year, month)[1]


def get_month_date_range(dt: datetime) -> tuple:
    """지정된 datetime 객체의 월 기준 시작일과 마지막 일을 "YYYY-MM-DD" 형태로
    반환한다. 원래 fetch_eloboard_data.py 안에 있던 함수였는데, 그 스크립트가
    HTML 스크래핑에서 EloBoard 공식 API(player_id 기반) 호출로 바뀌면서 날짜
    포맷도 "YYYYMMDD"(POST 파라미터용)에서 "YYYY-MM-DD"(API의 played_on과
    문자열 비교가 되는 ISO 포맷)로 함께 바뀌었다. update_data.py도 이 함수를
    가져다 쓰므로, 한쪽만 고치고 다른 쪽을 깜빡하는 일이 없도록 공용 모듈로
    옮겼다."""
    last_day = last_day_of_month(dt.year, dt.month)
    return dt.strftime("%Y-%m-01"), dt.strftime(f"%Y-%m-{last_day:02d}")


def fetch_json(url: str, *, method: str = "GET", params=None, data=None, headers=None,
                label: str = "", max_retries: int = HTTP_MAX_RETRIES,
                timeout: int = HTTP_TIMEOUT_SEC, backoff: int = HTTP_RETRY_BACKOFF_SEC):
    """JSON을 응답하는 API를 재시도와 함께 호출하는 공용 헬퍼.

    최대 max_retries번까지 시도하며, HTTP 오류/네트워크 예외/JSON 파싱 실패를
    전부 잡아서 경고 로그를 남기고 backoff초 뒤 재시도한다. 재시도를 다 소진하면
    예외를 던지는 대신 None을 반환한다 - 호출부가 "이번 요청은 실패, 기존 값
    유지"로 자연스럽게 처리할 수 있도록(부분 데이터를 정상 결과처럼 흘려보내지
    않기 위함이기도 하다).

    label은 로그 메시지 앞에 붙는 짧은 식별자(예: "풍고(2025-06, 30명분)",
    "엘로보드(offset=200)") - 여러 요청이 동시에/연달아 실패했을 때 로그에서
    어느 요청인지 구분하기 위함이다.
    """
    req_headers = {"User-Agent": USER_AGENT, "Accept": "application/json"}
    if headers:
        req_headers.update(headers)
    prefix = f"{label} " if label else ""

    for attempt in range(1, max_retries + 1):
        try:
            response = requests.request(
                method, url, params=params, data=data, headers=req_headers, timeout=timeout
            )
            if response.status_code != 200:
                print(f"[경고] {prefix}HTTP {response.status_code} 응답 ({attempt}/{max_retries})", file=sys.stderr)
            else:
                return response.json()
        except requests.RequestException as e:
            print(f"[경고] {prefix}요청 실패: {e} ({attempt}/{max_retries})", file=sys.stderr)
        except ValueError as e:
            # response.json() 파싱 실패 (JSONDecodeError는 ValueError의 하위클래스)
            print(f"[경고] {prefix}응답 파싱 실패: {e} ({attempt}/{max_retries})", file=sys.stderr)

        if attempt < max_retries:
            time.sleep(backoff)

    return None


def atomic_write_json(path: Path, data, **json_kwargs) -> None:
    """JSON 파일을 원자적으로(all-or-nothing) 쓴다.

    open(path, "w") + json.dump()를 직접 쓰면, 워크플로우가 쓰는 도중에
    죽거나(타임아웃, 강제취소, 러너 문제 등) 취소되면 중간까지만 써진 깨진
    JSON 파일이 그대로 남을 수 있다 - 그러면 다음 실행이 그 파일을 읽으려다
    파싱 에러로 죽고, 사람이 수동으로 고칠 때까지 계속 실패한다.

    대신 같은 디렉터리에 임시 파일로 먼저 전부 쓰고, 다 쓴 뒤에 os.replace()로
    한 번에 원래 이름으로 바꿔치기한다. os.replace()는 같은 파일시스템 안에서
    원자적 연산이라(운영체제가 보장), 중간 상태라는 게 존재하지 않는다 - 쓰다가
    죽어도 원래 파일은 쓰기 시작 전 그대로 안전하게 남는다.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    json_kwargs.setdefault("ensure_ascii", False)
    json_kwargs.setdefault("indent", 2)

    fd, tmp_path = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, **json_kwargs)
        os.replace(tmp_path, path)
    except Exception:
        # 실패하면 임시 파일 잔재를 남기지 않는다 - 원래 path는 애초에 안
        # 건드렸으므로 그대로 안전하다. KeyboardInterrupt/SystemExit까지
        # 여기서 잡을 필요는 없다(일반적인 쓰기 실패만 대상) - 어차피
        # tempfile.mkstemp가 매번 고유한 이름을 만들어서, 정리를 못 하고
        # 남는 임시파일이 있어도 다음 실행과 이름이 겹치거나 문제를 일으키지
        # 않는다.
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def safe_read_json(path: Path, default=None):
    """JSON 파일을 읽되, 파일이 없거나 깨져있으면 예외를 던지는 대신 default를
    반환한다. 상태 스냅샷 파일들(baseline, applied_corrections 등)처럼 "없으면
    처음부터 다시 시작"이 안전한 기본 동작인 곳에 쓴다 - 실수로 지워지거나
    (atomic_write_json을 쓰기 전 과거에) 깨진 채로 커밋된 파일 때문에 전체
    스크립트가 죽는 걸 막기 위함이다."""
    path = Path(path)
    if not path.exists():
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"[경고] {path}를 읽을 수 없어 기본값으로 대체합니다: {e}", file=sys.stderr)
        return default


REQUIRED_MEMBER_STRING_FIELDS = ("id", "nickname")
VALID_GENDERS = {"m", "f", None}


def validate_and_clean_members(members: list) -> list:
    """members.json의 각 레코드가 최소한의 스키마를 만족하는지 검사하고, 문제
    있는 레코드는 걸러내면서(고쳐서 넘어갈 수 있는 건 고쳐서) 경고를 남긴다.

    이 프로젝트는 xlsx/EloBoard API/admin.html처럼 서로 다른 소스가 전부
    members.json을 직접 건드릴 수 있는 구조라(오늘도 숫자 타입 id, 문자열
    "null" 같은 문제가 실제로 있었다), 다운스트림 코드가 "당연히 있겠거니"
    하고 m["nickname"]처럼 직접 접근하다 하나 깨진 레코드 때문에 스크립트
    전체가 죽는 일이 없도록, 쓰기 시작하는 지점에서 한 번 걸러준다.

    검사 항목:
      - id/nickname이 비어있는 레코드는 통째로 제외(이 프로젝트에서 사람을
        식별하는 최소 조건이라, 고쳐 쓸 방법이 없다)
      - id는 문자열로 강제 변환(숫자로 저장된 경우 대비)
      - elo_id는 존재하면 정수로 변환 시도, 실패하면 None으로 대체
      - gender가 "m"/"f"/None이 아니면 경고만 남기고 원래 값 유지(치명적이진
        않아서 걸러내지는 않음)
    """
    cleaned = []
    for idx, m in enumerate(members):
        if not isinstance(m, dict):
            print(f"[경고] members[{idx}]가 객체가 아님 - 건너뜀: {m!r}", file=sys.stderr)
            continue

        member_id = m.get("id")
        nickname = m.get("nickname")
        if not member_id or not nickname:
            print(f"[경고] members[{idx}]에 id 또는 nickname이 없어 건너뜀: {m!r}", file=sys.stderr)
            continue

        m = dict(m)  # 원본 리스트를 in-place로 안 건드리고 복사본을 정리한다
        m["id"] = str(member_id)

        elo_id = m.get("elo_id")
        if elo_id is not None:
            try:
                m["elo_id"] = int(elo_id)
            except (ValueError, TypeError):
                print(f"[경고] {m['id']}의 elo_id({elo_id!r})가 정수로 변환 안 돼 null로 대체", file=sys.stderr)
                m["elo_id"] = None

        if m.get("gender") not in VALID_GENDERS:
            print(f"[경고] {m['id']}의 gender 값이 예상과 다름: {m.get('gender')!r}", file=sys.stderr)

        cleaned.append(m)

    return cleaned


# ---------------------------------------------------------------------------
# xlsx 관련 공용 헬퍼
#
# data/members.xlsx가 로스터의 유일한 원본(source of truth)이다. 예전엔
# xlsx/members.json 양쪽에서 편집 가능해서 3-way 병합이 필요했는데, admin.html도
# 이제 xlsx를 직접 읽고 쓰도록 바뀌어서 편집 창구가 xlsx 하나로 통일됐다 - 그래서
# 여러 스크립트(convert_members_xlsx.py, sync_members.py, fetch_eloboard_data.py)가
# 전부 xlsx를 직접 읽거나 쓸 일이 있어서 여기 공용으로 모아둔다.
# ---------------------------------------------------------------------------

XLSX_PATH = ROOT / "data" / "members.xlsx"
XLSX_SHEET_NAME = "members"

XLSX_GENDER_MAP = {"남자": "m", "여자": "f"}
XLSX_GENDER_MAP_REVERSE = {"m": "남자", "f": "여자"}
XLSX_PLACEHOLDER_VALUES = {"체크", "todo", "?", "미정", "", "null", "none", "n/a", "na"}
XLSX_CORE_FIELDS = ["nickname", "elo_id", "birthdate", "gender", "race", "tier", "team", "role"]


def xlsx_clean(value):
    """"체크"/"null"류 임시 문자열이나 빈 문자열을 null로 정규화한다(대소문자
    구분 안 함). 그 외 값은 그대로 통과."""
    if isinstance(value, str) and value.strip().lower() in XLSX_PLACEHOLDER_VALUES:
        return None
    return value


def xlsx_format_date(value):
    value = xlsx_clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip() or None


def xlsx_parse_date_for_write(value):
    """members.json 스타일 "YYYY-MM-DD" 문자열을 엑셀에 넣을 datetime으로
    되돌린다. 형식이 안 맞거나 None이면 빈 칸(None)으로 남긴다."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def parse_xlsx_members(ws) -> dict:
    """워크시트를 읽어 soop_id -> {core_fields..., info_updated_at} 딕셔너리로
    반환한다. id/nickname이 비어있는 행은 건너뛴다."""
    rows = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nickname, soop_id, elo_id, birthdate, gender, race, tier, team, role, updated_at = row[:10]

        if not nickname or not soop_id:
            print(f"[건너뜀] {row_idx}행: 이름 또는 SOOP ID가 비어있음", file=sys.stderr)
            continue

        # SOOP ID가 순수 숫자면 엑셀/openpyxl이 문자열이 아니라 int로 읽어들인다.
        # 항상 문자열로 강제해서 이후 매칭/직렬화 문제를 원천 차단한다.
        soop_id = str(soop_id)
        tier = xlsx_clean(tier)

        rows[soop_id] = {
            "nickname": nickname,
            "elo_id": elo_id if elo_id is not None else None,
            "birthdate": xlsx_format_date(birthdate),
            "gender": XLSX_GENDER_MAP.get(gender, gender),
            "race": xlsx_clean(race),
            "tier": str(tier) if tier is not None else None,
            "team": xlsx_clean(team),
            "role": role if role else "",
            "info_updated_at": xlsx_format_date(updated_at),
        }
    return rows


def load_xlsx_members() -> dict:
    """XLSX_PATH를 열어 soop_id -> core_fields 딕셔너리로 반환한다. 파일이나
    시트가 없으면 빈 딕셔너리를 반환한다(에러로 죽지 않음 - 최초 실행 등에서
    xlsx 자체가 아직 없을 수 있음)."""
    if not XLSX_PATH.exists():
        return {}
    wb = load_workbook(XLSX_PATH, data_only=True)
    if XLSX_SHEET_NAME not in wb.sheetnames:
        print(f"[오류] '{XLSX_SHEET_NAME}' 시트를 찾을 수 없습니다. (있는 시트: {wb.sheetnames})", file=sys.stderr)
        return {}
    return parse_xlsx_members(wb[XLSX_SHEET_NAME])


def write_xlsx(update_rows: dict, append_rows: list, delete_ids: set | None = None) -> None:
    """update_rows: {soop_id: core_fields} - 기존 행을 이 값으로 갱신(수정일은 안 건드림).
    append_rows: [{"id":..., **core_fields, "info_updated_at":...}] - 새 행 추가.
    delete_ids: 이 id에 해당하는 행을 통째로 지운다."""
    delete_ids = delete_ids or set()
    wb = load_workbook(XLSX_PATH)  # data_only=False - 저장을 위해 다시 연다
    ws = wb[XLSX_SHEET_NAME]

    if update_rows or delete_ids:
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            cell_id = row[1].value
            cell_id = str(cell_id) if cell_id is not None else None
            if cell_id in delete_ids:
                rows_to_delete.append(row[0].row)
                continue
            fields = update_rows.get(cell_id)
            if not fields:
                continue
            row[0].value = fields.get("nickname")
            row[2].value = fields.get("elo_id")
            row[3].value = xlsx_parse_date_for_write(fields.get("birthdate"))
            row[4].value = XLSX_GENDER_MAP_REVERSE.get(fields.get("gender"), fields.get("gender"))
            row[5].value = fields.get("race")
            row[6].value = fields.get("tier")
            row[7].value = fields.get("team")
            row[8].value = fields.get("role") or None
            # "수정일"(row[9])은 여기서 건드리지 않는다 - 아카이브 소급 정정
            # 전용 필드라, 자동 갱신 스크립트가 값을 넣으면 의도치 않게 소급
            # 정정이 걸려버린다.

        # 뒤에서부터(행 번호가 큰 것부터) 지워야, 먼저 지운 행 때문에 아직
        # 안 지운 행들의 번호가 밀리는 문제가 안 생긴다.
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    for m in append_rows:
        ws.append([
            m.get("nickname"),
            m.get("id"),
            m.get("elo_id"),
            xlsx_parse_date_for_write(m.get("birthdate")),
            XLSX_GENDER_MAP_REVERSE.get(m.get("gender"), m.get("gender")),
            m.get("race"),
            m.get("tier"),
            m.get("team"),
            m.get("role") or None,
            xlsx_parse_date_for_write(m.get("info_updated_at")),
        ])

    wb.save(XLSX_PATH)
