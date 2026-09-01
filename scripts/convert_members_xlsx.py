"""
data/members.xlsx (사람이 직접 관리하는 엑셀 명단)와 data/members.json을 진짜
양방향으로("어디서 고치든 최신 게 반영") 동기화하는 스크립트.

핵심 문제: xlsx와 json 둘 다 사람이 직접 고칠 수 있는데(xlsx는 엑셀로, json은
admin.html로), 둘 다 "값이 있음"이라 어느 쪽이 최신인지 값만 보고는 판단할 수
없다. info_updated_at("수정일")은 이 판단에 쓸 수 없다 - 그건 아카이브 소급
정정 전용 필드라, 동기화 승자 판단에 갖다 쓰면 그때마다 의도치 않게 과거
아카이브까지 소급 수정돼버린다.

그래서 "지난 성공적인 동기화 직후의 상태"를 스냅샷(data/members_sync_baseline.json)
으로 저장해두고, 이번 실행에서 xlsx/json 각각을 그 스냅샷과 비교해서 "어느 쪽이
그 사이에 바뀌었는지"를 3-way diff로 판단한다:
  - xlsx만 바뀜 -> xlsx 값을 json에 반영
  - json만 바뀜(예: admin.html로 수정) -> json 값을 xlsx에 반영
  - 둘 다 바뀌었고 서로 다름(충돌) -> xlsx를 기본으로 채택하되 [경고]로 명확히
    알린다(사람이 확인하고 필요하면 직접 정리하도록)
  - 둘 다 안 바뀜 -> 아무것도 안 함
동기화 대상은 nickname/elo_id/birthdate/gender/race/tier/team/role 8개 필드다.

info_updated_at("수정일")은 이 3-way 병합과 완전히 별개로, 기존 규칙 그대로
유지한다: xlsx의 "수정일" 셀이 채워져 있으면 그 값으로 json을 갱신하고, 비어있으면
기존 json 값을 그대로 보존한다(admin.html에서 넣어둔 값이 매일 자동 실행 때마다
사라지지 않도록).

신규/삭제 처리:
  - xlsx에만 있는 사람(신규) -> json에 추가
  - json에만 있는 사람(신규) -> xlsx에 새 행으로 추가
  - 스냅샷엔 있었는데 한쪽에서 사라진 사람 -> 삭제로 취급하지 않는다(이 프로젝트
    전반의 정책과 동일 - 자동으로 사람을 지우지 않음). 어느 한쪽에 남아있는
    정보를 기준으로 다시 채워질 수 있다.

엑셀 시트("members") 컬럼 매핑:
  이름 <-> nickname, SOOP ID <-> id, ELO ID <-> elo_id, 생년월일 <-> birthdate,
  성별(남자/여자) <-> gender(m/f), 종족 <-> race, 티어 <-> tier(문자열로 통일),
  소속 <-> team, 직책 <-> role(비어있으면 ""), 수정일 <-> info_updated_at

"체크"/"null"처럼 확인 전 임시로 박아둔 문자열은 xlsx -> json 방향에서 null로
정규화한다(대소문자 구분 없음). json -> xlsx 방향에서는 반대로 None 값을 빈
칸으로 남긴다.

xlsx 파일 자체가 없으면 두 방향 다 에러로 죽지 않고 조용히 건너뛴다.

실행:
  python scripts/convert_members_xlsx.py --dry-run   (미리보기, 파일 안 바꿈)
  python scripts/convert_members_xlsx.py             (실제 반영)
"""

import sys
import os
import json
import argparse
import tempfile
from datetime import datetime

from openpyxl import load_workbook

from _common import ROOT, atomic_write_json, safe_read_json, validate_and_clean_members

XLSX_PATH = ROOT / "data" / "members.xlsx"
MEMBERS_PATH = ROOT / "data" / "members.json"
BASELINE_PATH = ROOT / "data" / "members_sync_baseline.json"
SHEET_NAME = "members"

CORE_FIELDS = ["nickname", "elo_id", "birthdate", "gender", "race", "tier", "team", "role"]

GENDER_MAP = {"남자": "m", "여자": "f"}
GENDER_MAP_REVERSE = {"m": "남자", "f": "여자"}
PLACEHOLDER_VALUES = {"체크", "todo", "?", "미정", "", "null", "none", "n/a", "na"}


def clean(value):
    """"체크"/"null"류 임시 문자열이나 빈 문자열을 null로 정규화한다(대소문자
    구분 안 함). 그 외 값은 그대로 통과."""
    if isinstance(value, str) and value.strip().lower() in PLACEHOLDER_VALUES:
        return None
    return value


def format_date(value):
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip() or None


def parse_date_for_xlsx(value):
    """members.json의 "YYYY-MM-DD" 문자열을 엑셀에 넣을 datetime으로 되돌린다.
    형식이 안 맞거나 None이면 빈 칸(None)으로 남긴다."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def parse_xlsx_members(ws) -> dict:
    """id -> {core_fields..., info_updated_at} 딕셔너리로 반환."""
    rows = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nickname, soop_id, elo_id, birthdate, gender, race, tier, team, role, updated_at = row[:10]

        if not nickname or not soop_id:
            print(f"[건너뜀] {row_idx}행: 이름 또는 SOOP ID가 비어있음", file=sys.stderr)
            continue

        # SOOP ID가 순수 숫자면 엑셀/openpyxl이 문자열이 아니라 int로 읽어들인다.
        # 항상 문자열로 강제해서 이후 매칭/직렬화 문제를 원천 차단한다.
        soop_id = str(soop_id)
        tier = clean(tier)

        rows[soop_id] = {
            "nickname": nickname,
            "elo_id": elo_id if elo_id is not None else None,
            "birthdate": format_date(birthdate),
            "gender": GENDER_MAP.get(gender, gender),
            "race": clean(race),
            "tier": str(tier) if tier is not None else None,
            "team": clean(team),
            "role": role if role else "",
            "info_updated_at": format_date(updated_at),
        }
    return rows


def core(fields: dict) -> dict:
    return {k: fields.get(k) for k in CORE_FIELDS}


def write_xlsx(update_rows: dict, append_rows: list, delete_ids: set | None = None) -> None:
    """update_rows: {soop_id: core_fields} - 기존 행을 이 값으로 갱신(수정일은 안 건드림).
    append_rows: [{"id":..., **core_fields, "info_updated_at":...}] - 새 행 추가.
    delete_ids: 이 id에 해당하는 행을 통째로 지운다 - admin.html에서 사람의
    id가 바뀌어서(예: "미상(elo_N)" -> 실제 아이디) xlsx에만 예전 id로 남아있게
    된 오래된 행을 정리할 때 쓴다(안 지우면 매번 다시 json에 되살아나는
    버그가 있었음)."""
    delete_ids = delete_ids or set()
    wb = load_workbook(XLSX_PATH)  # data_only=False - 저장을 위해 다시 연다
    ws = wb[SHEET_NAME]

    if update_rows or delete_ids:
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            cell_id = row[1].value
            cell_id = str(cell_id) if cell_id is not None else None
            if cell_id in delete_ids:
                rows_to_delete.append(row[0].row)
                continue
            fields = update_rows.get(cell_id)
            if not fields:
                continue
            row[0].value = fields.get("nickname")
            row[2].value = fields.get("elo_id")
            row[3].value = parse_date_for_xlsx(fields.get("birthdate"))
            row[4].value = GENDER_MAP_REVERSE.get(fields.get("gender"), fields.get("gender"))
            row[5].value = fields.get("race")
            row[6].value = fields.get("tier")
            row[7].value = fields.get("team")
            row[8].value = fields.get("role") or None
            # "수정일"(row[9])은 이 3-way 병합과 무관한 필드라 여기서 건드리지 않는다.

        # 뒤에서부터(행 번호가 큰 것부터) 지워야, 먼저 지운 행 때문에 아직
        # 안 지운 행들의 번호가 밀리는 문제가 안 생긴다.
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

    for m in append_rows:
        ws.append((
            m.get("nickname"),
            m.get("id"),
            m.get("elo_id"),
            parse_date_for_xlsx(m.get("birthdate")),
            GENDER_MAP_REVERSE.get(m.get("gender"), m.get("gender")),
            m.get("race"),
            m.get("tier"),
            m.get("team"),
            m.get("role") or None,
            parse_date_for_xlsx(m.get("info_updated_at")),
        ))

    # wb.save()도 json.dump()와 똑같이 대상 파일에 직접 쓰는 방식이라, 쓰는
    # 도중에 중단되면 xlsx 파일 자체가 깨질 수 있다(엑셀에서 열리지도 않는
    # 상태) - 임시 파일에 먼저 저장하고 다 되면 os.replace()로 바꿔치기한다.
    fd, tmp_path = tempfile.mkstemp(dir=XLSX_PATH.parent, prefix=f".{XLSX_PATH.name}.", suffix=".tmp")
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, XLSX_PATH)
    except BaseException:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def resolve_conflict(xlsx_fields: dict, json_fields: dict, base_fields: dict | None) -> tuple:
    """xlsx/json 3-way 병합의 승자 판단 로직. base_fields는 지난 동기화 스냅샷
    (없으면 None - 첫 실행이라 하위 호환으로 xlsx를 기본 채택).
    반환값: (winner_fields, is_conflict)."""
    xlsx_changed = base_fields is None or xlsx_fields != base_fields
    json_changed = base_fields is not None and json_fields != base_fields

    if xlsx_changed and json_changed and xlsx_fields != json_fields:
        return xlsx_fields, True  # 충돌 시 기본값: xlsx
    if json_changed and not xlsx_changed:
        return json_fields, False
    return xlsx_fields, False


def main(argv=None):
    """argv=None이면(평소 CLI 실행) sys.argv[1:]를 읽는다. pytest 같은 테스트
    러너가 자기 옵션(-v 등)을 sys.argv에 남겨둔 채로 main()을 직접 호출하면
    argparse가 그걸 이 스크립트의 옵션으로 착각해서 죽는데, 테스트에서
    main(argv=[])처럼 명시적으로 넘겨주면 그 문제가 안 생긴다."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="파일을 실제로 바꾸지 않고 결과만 출력")
    args = parser.parse_args(argv)

    if not XLSX_PATH.exists():
        print(f"[건너뜀] {XLSX_PATH} 가 없습니다 - members.json은 그대로 둡니다.", file=sys.stderr)
        return

    wb = load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[오류] '{SHEET_NAME}' 시트를 찾을 수 없습니다. (있는 시트: {wb.sheetnames})", file=sys.stderr)
        sys.exit(1)

    xlsx_members = parse_xlsx_members(wb[SHEET_NAME])
    print(f"[준비] {XLSX_PATH.name}에서 {len(xlsx_members)}명 확인")

    local_data = {"members": []}
    if MEMBERS_PATH.exists():
        local_data = safe_read_json(MEMBERS_PATH, default={"members": []})

    # id 숫자 타입 정규화, elo_id 정수 강제, id/nickname 없는 레코드 제외 등을
    # 여기서 한 번에 처리한다 - convert_members_xlsx.py가 워크플로우에서 제일
    # 먼저 members.json을 만지는 스크립트라, 여기서 걸러두면 뒤따르는
    # sync_members.py/update_data.py는 항상 정리된 데이터를 받는다.
    before_count = len(local_data.get("members", []))
    local_data["members"] = validate_and_clean_members(local_data.get("members", []))
    dropped = before_count - len(local_data["members"])
    if dropped:
        print(f"[정리] members.json에서 스키마 문제로 {dropped}명 제외됨", file=sys.stderr)

    member_map = {m["id"]: m for m in local_data["members"]}

    # elo_id -> 현재 json에서의 id. admin.html에서 "미상(elo_N)" 임시 프로필의
    # id를 실제 SOOP 아이디로 바꿔치기하면(nickname/team 등도 같이 채움), 그
    # 사람은 json에서 새 id로 존재하게 되는데 - xlsx는 여전히 예전 id(elo_N)로
    # 그 사람을 기억하고 있다(admin.html은 xlsx를 안 건드리므로). 이 룩업이
    # 있어야 "xlsx에는 있는데 json엔 없는" 경우를 만났을 때, 그게 진짜 삭제된
    # 사람인지 아니면 그냥 id가 바뀐 사람인지 구분할 수 있다.
    def _norm_elo_id(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    json_elo_id_to_id = {
        eid: sid for sid, m in member_map.items()
        if (eid := _norm_elo_id(m.get("elo_id"))) is not None
    }

    baseline = safe_read_json(BASELINE_PATH, default={})

    all_ids = set(xlsx_members) | set(member_map)

    json_updates, xlsx_updates, added_to_json, added_to_xlsx, conflicts = [], {}, [], [], []
    stale_xlsx_ids = []
    new_baseline = {}

    for soop_id in all_ids:
        in_xlsx = soop_id in xlsx_members
        in_json = soop_id in member_map

        if in_xlsx and in_json:
            xlsx_fields = core(xlsx_members[soop_id])
            json_fields = core(member_map[soop_id])
            base_fields = baseline.get(soop_id)

            winner, is_conflict = resolve_conflict(xlsx_fields, json_fields, base_fields)
            if is_conflict:
                conflicts.append((soop_id, xlsx_fields, json_fields))

            if json_fields != winner:
                member_map[soop_id].update(winner)
                json_updates.append((soop_id, json_fields, winner))
            if xlsx_fields != winner:
                xlsx_updates[soop_id] = winner

            # info_updated_at은 3-way 병합과 무관하게 기존 규칙 그대로: xlsx
            # "수정일"이 채워져 있을 때만 반영, 비어있으면 기존 json 값 보존.
            xlsx_info_updated_at = xlsx_members[soop_id]["info_updated_at"]
            if xlsx_info_updated_at is not None:
                member_map[soop_id]["info_updated_at"] = xlsx_info_updated_at

            new_baseline[soop_id] = winner

        elif in_xlsx and not in_json:
            xlsx_elo_id = _norm_elo_id(xlsx_members[soop_id].get("elo_id"))
            renamed_to = json_elo_id_to_id.get(xlsx_elo_id) if xlsx_elo_id is not None else None
            if renamed_to is not None:
                # 이 xlsx 행은 admin.html 등에서 id가 바뀐 사람의 예전 흔적이다
                # (같은 elo_id를 가진 진짜 레코드가 이미 json에 다른 id로 존재).
                # 예전처럼 이걸 json에 다시 만들어내면(= "미상" 중복 생성 버그),
                # 매번 이 검사를 할 때마다 계속 되살아난다 - 그래서 json에
                # 되살리지 않고, 대신 xlsx 쪽의 이 오래된 행 자체를 지운다.
                stale_xlsx_ids.append(soop_id)
                continue

            fields = core(xlsx_members[soop_id])
            new_member = dict(fields)
            new_member["id"] = soop_id
            new_member["info_updated_at"] = xlsx_members[soop_id]["info_updated_at"]
            local_data.setdefault("members", []).append(new_member)
            member_map[soop_id] = new_member
            added_to_json.append((soop_id, fields["nickname"]))
            new_baseline[soop_id] = fields

        elif in_json and not in_xlsx:
            fields = core(member_map[soop_id])
            append_obj = dict(fields)
            append_obj["id"] = soop_id
            append_obj["info_updated_at"] = member_map[soop_id].get("info_updated_at")
            added_to_xlsx.append(append_obj)
            new_baseline[soop_id] = fields
        # 스냅샷엔 있었는데 이번엔 둘 다에서 사라진 경우 -> new_baseline에서 자연히
        # 빠짐(더 이상 추적 안 함). 삭제로 취급해 어느 파일도 건드리지 않는다.

    if args.dry_run:
        for soop_id, before, after in json_updates:
            print(f"  [json 갱신 예정] {soop_id}: {before} -> {after}")
        for soop_id, fields in xlsx_updates.items():
            print(f"  [xlsx 갱신 예정] {soop_id}: -> {fields}")
        for soop_id, nickname in added_to_json:
            print(f"  [json 추가 예정] {soop_id} ({nickname})")
        for m in added_to_xlsx:
            print(f"  [xlsx 추가 예정] {m['id']} ({m['nickname']})")
        for soop_id in stale_xlsx_ids:
            print(f"  [xlsx 삭제 예정] {soop_id} (id가 바뀌어 json엔 이미 다른 id로 존재 - 예전 흔적 정리)")
        for soop_id, xlsx_fields, json_fields in conflicts:
            print(f"  [충돌!] {soop_id}: xlsx={xlsx_fields} vs json={json_fields} -> xlsx 값을 기본 채택")
        print(f"[dry-run 완료] json 갱신 {len(json_updates)} / xlsx 갱신 {len(xlsx_updates)} / "
              f"json 추가 {len(added_to_json)} / xlsx 추가 {len(added_to_xlsx)} / xlsx 삭제 {len(stale_xlsx_ids)} / "
              f"충돌 {len(conflicts)}건 (파일은 안 건드림)")
        return

    for soop_id, xlsx_fields, json_fields in conflicts:
        print(f"[경고] {soop_id}: xlsx와 json이 동시에 바뀌어 충돌 - xlsx 값을 기본 채택함. "
              f"xlsx={xlsx_fields} / json(무시됨)={json_fields}", file=sys.stderr)

    atomic_write_json(MEMBERS_PATH, local_data)
    print(f"[완료] members.json 동기화됨 (갱신 {len(json_updates)}명, 신규 {len(added_to_json)}명, "
          f"총 {len(local_data['members'])}명)")

    if xlsx_updates or added_to_xlsx or stale_xlsx_ids:
        write_xlsx(xlsx_updates, added_to_xlsx, delete_ids=set(stale_xlsx_ids))
        print(f"[완료] members.xlsx 동기화됨 (갱신 {len(xlsx_updates)}명, 신규 추가 {len(added_to_xlsx)}명, "
              f"오래된 행 삭제 {len(stale_xlsx_ids)}명)")
    else:
        print("[완료] xlsx 쪽 변경 없음")

    atomic_write_json(BASELINE_PATH, new_baseline)


if __name__ == "__main__":
    main()
